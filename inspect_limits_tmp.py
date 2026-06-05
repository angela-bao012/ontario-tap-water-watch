import openpyxl
import os

DATA_DIR = "/Users/angelabao/ontario-tap-water-watch/data"
FILES = [
    'Test Data - Raw Data 2016 17.xlsx',
    'Test Data - Raw Data 2017_18.xlsx',
    'Test Data - Raw Data 2018_19.xlsx',
    'Test Data - Raw Data 2019_20_EN.xlsx',
    'Test Data - Raw Data 2020_2021.xlsx',
    'Test Results 2021-22 EN.xlsx',
    'Test Results 2022-23 EN.xlsx',
    'Test Results 2023-24 EN .xlsx',
    'Test Results 2024-25 EN.xlsx',
]

def inspect_all():
    lead_limits = set()
    for filename in FILES:
        filepath = os.path.join(DATA_DIR, filename)
        if not os.path.exists(filepath):
            continue
        print(f"Inspecting {filename}...")
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        
        try:
            param_idx = header_row.index('Parameter Name')
        except ValueError:
            continue
            
        limit_idx = None
        for col_name in ['Parameter Limit', 'Limit']:
            if col_name in header_row:
                limit_idx = header_row.index(col_name)
                break
                
        if limit_idx is None:
            continue
            
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            param_val = row[param_idx]
            limit_val = row[limit_idx]
            if param_val and 'lead' in str(param_val).lower():
                lead_limits.add(limit_val)
            row_count += 1
            if row_count > 5000:
                break
        wb.close()
    print(f"All unique Lead limits across files: {lead_limits}")

if __name__ == '__main__':
    inspect_all()
