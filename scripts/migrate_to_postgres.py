#!/usr/bin/env python3
"""
Migrate Ontario drinking water testing Excel files into a local PostgreSQL database.
Directly ingests original, raw system names without any renaming/cleaning logic,
and maps all metadata (owner_name, sample_type, parameter_limit) in a single pass.
"""

import openpyxl
import psycopg2
from psycopg2.extras import execute_values
import os
import re
from datetime import datetime

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data')
DB_CONFIG = {
    "host": "localhost",
    "port": 5432,
    "user": "postgres",
    "password": "68391975",
    "dbname": "water_watch"
}

FILES = [
    'Test Data - Raw Data 2016 17.xlsx',
    'Test Data - Raw Data 2017_18.xlsx',
    'Test Data - Raw Data 2018_19.xlsx',
    'Test Data - Raw Data 2019_20_EN.xlsx',
    'Test Data - Raw Data 2020_2021.xlsx',
    'Test Results 2021-22 EN.xlsx',
    'Test Results 2022-23 EN.xlsx',
    'Test Results 2023-24 EN .xlsx',
    'Test Results 2024-25 EN.xlsx',
]

def normalize_header(name):
    if name is None:
        return None
    name = str(name).strip()
    mapping = {
        'DWS #': 'dws_id',
        'DWS Name': 'dws_name',
        'Owner Legal Name': 'owner_name',
        'Regulation Name': 'regulation',
        'Regulation': 'regulation',
        'DWS Category': 'dws_category',
        'PHU Legal Name': 'phu_name',
        'Sample Date': 'sample_date',
        'Sample Type Name': 'sample_type',
        'Sample Location': 'sample_location',
        'Parameter Name': 'parameter_name',
        'Result': 'result_value',
        'Result Units': 'result_unit',
        'Parameter Limit': 'parameter_limit',
        'Exceedance': 'exceedance',
        'Parameter Group': 'parameter_group',
        'DWIS Result Record ID': 'dwis_record_id',
        'DWIS Result Record Id': 'dwis_record_id',
        'Qualifier': 'qualifier',
    }
    return mapping.get(name, name.lower().replace(' ', '_'))


def parse_result(val):
    if val is None or val == '' or val == '-' or val == 'N/A':
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(',', '').strip())
    except (ValueError, AttributeError):
        return None


def get_date(sample_date):
    """Parse sample date into a date object, returns None if not possible."""
    if sample_date is None:
        return None
    if isinstance(sample_date, datetime):
        return sample_date.date()
    if hasattr(sample_date, 'date'):
        return sample_date.date()
    if isinstance(sample_date, str):
        s = sample_date.strip()
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%Y/%m/%d', '%d-%m-%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S'):
            try:
                return datetime.strptime(s[:10], fmt).date()
            except ValueError:
                continue
        # Fallback: find 4-digit year
        m = re.search(r'(20\d{2})', s)
        if m:
            # return Jan 1st of that year if only year can be found
            return datetime(int(m.group(1)), 1, 1).date()
    return None


def clean_location(raw_name):
    # Completely eliminate renaming/cleaning logic, preserve exactly as is
    if not raw_name:
        return "Unknown Location"
    return str(raw_name).strip()


def get_unique_cleaned_name(dws_id, dws_name):
    # Completely eliminate renaming/cleaning logic, preserve exactly as is
    if not dws_name:
        return "Unknown Location"
    return str(dws_name).strip()


def title_case_name(s):
    if not s:
        return ""
    words = s.split()
    title_words = []
    
    lowercase_words = {'of', 'and', 'the', 'for', 'in', 'on', 'at', 'to', 'by', 'with', 'a', 'an'}
    
    for i, w in enumerate(words):
        clean_w = re.sub(r'[^\w]', '', w).lower()
        if clean_w in lowercase_words and i > 0:
            title_words.append(w.lower())
        else:
            if '-' in w:
                sub_parts = w.split('-')
                sub_parts_title = [sub_p.capitalize() if len(sub_p) > 0 else "" for sub_p in sub_parts]
                title_words.append('-'.join(sub_parts_title))
            else:
                if len(w) > 1 and w[0].isalpha() and w[1] == '.':
                    title_words.append(w[0].upper() + w[1:])
                else:
                    title_words.append(w.capitalize())
                    
    return ' '.join(title_words)


def clean_infra(text):
    infra_words = [
        'DRINKING WATER SYSTEM', 'DRINKING WATER', 'WATER SYSTEM', 'DISTRIBUTION SYSTEM', 
        'DISTRIBUTION', 'WELL SUPPLY', 'SUPPLY WELL', 'SUPPLY', 'FACILITY'
    ]
    cleaned = text
    for w in infra_words:
        cleaned = re.sub(r'\b' + re.escape(w) + r'\b', '', cleaned, flags=re.IGNORECASE)
    return ' '.join(cleaned.split()).strip()


def make_display_name(raw_name):
    if not raw_name:
        return ""
    
    s = re.sub(r'^(R\d{3,6}\s*-\s*)|^(R\d{3,6}\s+)', '', raw_name, flags=re.IGNORECASE)
    s = re.sub(r'\s*\([#\d]+\)\s*$', '', s)
    s = ' '.join(s.split()).strip()
    
    s = re.sub(r'\bWELL SUPPLY\b', 'WELL', s, flags=re.IGNORECASE)
    s = re.sub(r'\bSUPPLY WELL\b', 'WELL', s, flags=re.IGNORECASE)
    
    if re.search(r'\s+-\s+', s):
        parts = [p.strip() for p in re.split(r'\s+-\s+', s)]
        if len(parts) >= 2:
            region_idx = -1
            region_indicators = ["CITY OF", "TOWN OF", "TOWNSHIP OF", "MUNICIPALITY OF", "REGIONAL MUNICIPALITY", "COUNTY OF", "DISTRICT OF", "COMMUNITY OF"]
            for idx, part in enumerate(parts):
                if any(ind in part.upper() for ind in region_indicators):
                    region_idx = idx
                    break
            
            if region_idx != -1:
                region_raw = parts[region_idx]
                other_parts = [p for i, p in enumerate(parts) if i != region_idx]
                facility_raw = " - ".join(other_parts)
            else:
                region_raw = parts[0]
                facility_raw = " - ".join(parts[1:])
            
            region_clean = clean_infra(region_raw)
            facility_clean = clean_infra(facility_raw)
            
            region_title = title_case_name(region_clean)
            facility_title = title_case_name(facility_clean)
            
            facility_title = re.sub(r'\b([A-Z])\.\s+([A-Z])\.', r'\1.\2.', facility_title, flags=re.IGNORECASE)
            
            if any(term in region_raw.upper() for term in ["DRINKING WATER SYSTEM", "WATER TREATMENT", "WTP"]):
                if not any(facility_title.lower().endswith(term) for term in ["plant", "well", "system", "school", "centre"]):
                    facility_title = f"{facility_title} Water Plant"
            
            return f"{facility_title} - {region_title}"

    cleaned = clean_infra(s)
    title_name = title_case_name(cleaned)
    return title_name


def setup_postgres_table():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    
    print("Dropping existing table if present...")
    cur.execute("DROP TABLE IF EXISTS water_quality_data;")
    
    print("Creating new table water_quality_data...")
    cur.execute("""
        CREATE TABLE water_quality_data (
            id SERIAL PRIMARY KEY,
            location VARCHAR(255) NOT NULL,
            sample_date DATE,
            parameter_name VARCHAR(255) NOT NULL,
            result_value DOUBLE PRECISION,
            result_unit VARCHAR(50),
            exceedance VARCHAR(10),
            owner_name VARCHAR(255),
            sample_type VARCHAR(100),
            parameter_limit VARCHAR(50),
            display_name VARCHAR(255)
        );
    """)
    
    print("Creating indexes...")
    cur.execute("CREATE INDEX idx_water_data_location ON water_quality_data(location);")
    cur.execute("CREATE INDEX idx_water_data_parameter ON water_quality_data(parameter_name);")
    cur.execute("CREATE INDEX idx_water_data_date ON water_quality_data(sample_date);")
    
    conn.commit()
    cur.close()
    conn.close()
    print("Table setup completed.")


def migrate_data():
    setup_postgres_table()
    
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    
    total_inserted = 0
    
    for filename in FILES:
        filepath = os.path.join(DATA_DIR, filename)
        if not os.path.exists(filepath):
            print(f"  SKIPPING: {filename} (not found)")
            continue
            
        print(f"Processing: {filename}...")
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [normalize_header(h) for h in header_row]
        
        if 'dws_id' not in headers or 'parameter_name' not in headers:
            print(f"  WARNING: {filename} has unexpected headers, skipping")
            wb.close()
            continue
            
        batch = []
        row_count = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i] is not None:
                    row_dict[headers[i]] = val
            
            dws_id = row_dict.get('dws_id')
            dws_name = row_dict.get('dws_name')
            parameter = row_dict.get('parameter_name')
            
            if parameter is None:
                continue
                
            # Get original raw location name exactly as written in raw file
            location = get_unique_cleaned_name(dws_id, dws_name)
            
            # Format and parse sample date
            s_date = get_date(row_dict.get('sample_date'))
            
            # Parse result value
            r_value = parse_result(row_dict.get('result_value'))
            
            # Exceedance formatting
            exceed = row_dict.get('exceedance')
            exceed_str = str(exceed).strip().upper() if exceed is not None else None
            if exceed_str not in ('Y', 'N'):
                exceed_str = None
                
            owner = str(row_dict.get('owner_name') or '').strip()
            s_type = str(row_dict.get('sample_type') or '').strip()
            limit = str(row_dict.get('parameter_limit') or '').strip()
            
            disp_name = make_display_name(location)
            
            batch.append((
                location,
                s_date,
                str(parameter),
                r_value,
                str(row_dict.get('result_unit') or ''),
                exceed_str,
                owner if owner else None,
                s_type if s_type else None,
                limit if limit else None,
                disp_name if disp_name else None
            ))
            row_count += 1
            
            # Batch insert in chunks of 10,000 to balance speed and memory
            if len(batch) >= 10000:
                execute_values(cur, """
                    INSERT INTO water_quality_data 
                    (location, sample_date, parameter_name, result_value, result_unit, exceedance, owner_name, sample_type, parameter_limit, display_name)
                    VALUES %s
                """, batch)
                conn.commit()
                total_inserted += len(batch)
                batch = []
                print(f"    Inserted {total_inserted:,} rows so far...", end='\r')
                
        if batch:
            execute_values(cur, """
                INSERT INTO water_quality_data 
                (location, sample_date, parameter_name, result_value, result_unit, exceedance, owner_name, sample_type, parameter_limit, display_name)
                VALUES %s
            """, batch)
            conn.commit()
            total_inserted += len(batch)
            
        print(f"    Finished {filename}: {row_count:,} rows parsed. Total database size: {total_inserted:,} rows.")
        wb.close()
        
    # Final database verification
    cur.execute("SELECT COUNT(*) FROM water_quality_data;")
    db_count = cur.fetchone()[0]
    
    cur.execute("SELECT COUNT(DISTINCT location) FROM water_quality_data;")
    loc_count = cur.fetchone()[0]
    
    print(f"\nMigration complete! Total rows in Postgres: {db_count:,}")
    print(f"Total unique locations: {loc_count:,}")
    
    # Confirmation logs of first 5 raw unique locations
    cur.execute("SELECT DISTINCT location FROM water_quality_data ORDER BY location LIMIT 5;")
    preview_locs = cur.fetchall()
    print("\n[✓] Raw Locations Ingested Preview:")
    for idx, l in enumerate(preview_locs, 1):
        print(f"  {idx}. {l[0]}")
        
    cur.close()
    conn.close()


if __name__ == '__main__':
    print("=" * 60)
    print("Migrating Ontario Water Data to PostgreSQL (Raw Naming)")
    print("=" * 60)
    migrate_data()
