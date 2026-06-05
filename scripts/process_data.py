#!/usr/bin/env python3
"""
Process Ontario drinking water testing Excel files into a SQLite database.

Reads all 9 Excel files (2016-2025), extracts year from actual sample dates,
and creates a queryable SQLite database with locations and test results.
"""

import openpyxl
import sqlite3
import os
import re
from datetime import datetime

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data')
DB_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'water_data.db')

FILES = [
    'Test Data - Raw Data 2016 17.xlsx',
    'Test Data - Raw Data 2017_18.xlsx',
    'Test Data - Raw Data 2018_19.xlsx',
    'Test Data - Raw Data 2019_20_EN.xlsx',
    'Test Data - Raw Data 2020_2021.xlsx',
    'Test Results 2021-22 EN.xlsx',
    'Test Results 2022-23 EN.xlsx',
    'Test Results 2023-24 EN .xlsx',
    'Test Results 2024-25 EN.xlsx',
]

def normalize_header(name):
    if name is None:
        return None
    name = name.strip()
    mapping = {
        'DWS #': 'dws_id',
        'DWS Name': 'dws_name',
        'Owner Legal Name': 'owner_name',
        'Regulation Name': 'regulation',
        'Regulation': 'regulation',
        'DWS Category': 'dws_category',
        'PHU Legal Name': 'phu_name',
        'Sample Date': 'sample_date',
        'Sample Type Name': 'sample_type',
        'Sample Location': 'sample_location',
        'Parameter Name': 'parameter_name',
        'Result': 'result_value',
        'Result Units': 'result_unit',
        'Parameter Limit': 'parameter_limit',
        'Exceedance': 'exceedance',
        'Parameter Group': 'parameter_group',
        'DWIS Result Record ID': 'dwis_record_id',
        'DWIS Result Record Id': 'dwis_record_id',
        'Qualifier': 'qualifier',
    }
    return mapping.get(name, name.lower().replace(' ', '_'))


def parse_result(val):
    if val is None or val == '' or val == '-' or val == 'N/A':
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(',', '').strip())
    except (ValueError, AttributeError):
        return None


def get_year(sample_date):
    """Extract year from sample date, returns None if not possible."""
    if sample_date is None:
        return None
    if hasattr(sample_date, 'year'):
        return sample_date.year
    if isinstance(sample_date, str):
        s = sample_date.strip()
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%Y/%m/%d', '%d-%m-%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S'):
            try:
                return datetime.strptime(s[:10], fmt).year
            except ValueError:
                continue
        # Fallback: find 4-digit year
        m = re.search(r'(20\d{2})', s)
        if m:
            return int(m.group(1))
    return None


def build_database():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.execute("PRAGMA journal_mode=WAL")
    cur.execute("PRAGMA synchronous=OFF")

    # Create tables (drop and recreate for clean rebuild)
    cur.execute("DROP TABLE IF EXISTS test_results")
    cur.execute("DROP TABLE IF EXISTS locations")
    cur.execute("DROP TABLE IF EXISTS locations_fts")
    conn.commit()

    cur.execute("""
        CREATE TABLE locations (
            dws_id INTEGER PRIMARY KEY,
            dws_name TEXT NOT NULL,
            owner_name TEXT,
            regulation TEXT,
            dws_category TEXT,
            phu_name TEXT,
            city_hint TEXT
        )
    """)
    
    cur.execute("""
        CREATE TABLE test_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dws_id INTEGER NOT NULL,
            sample_date TEXT,
            sample_type TEXT,
            sample_location TEXT,
            parameter_name TEXT NOT NULL,
            result_value REAL,
            result_unit TEXT,
            parameter_limit TEXT,
            exceedance TEXT,
            parameter_group TEXT,
            dwis_record_id INTEGER,
            qualifier TEXT,
            year INTEGER
        )
    """)

    cur.execute("CREATE INDEX idx_tests_dws_id ON test_results(dws_id)")
    cur.execute("CREATE INDEX idx_tests_parameter ON test_results(parameter_name)")
    cur.execute("CREATE INDEX idx_tests_year ON test_results(year)")
    cur.execute("CREATE INDEX idx_locations_name ON locations(dws_name)")
    cur.execute("CREATE INDEX idx_locations_phu ON locations(phu_name)")
    conn.commit()

    total_rows = 0
    locations_set = {}

    for filename in FILES:
        filepath = os.path.join(DATA_DIR, filename)
        if not os.path.exists(filepath):
            print(f"  SKIPPING: {filename} (not found)")
            continue
        
        print(f"Processing: {filename}...")
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [normalize_header(h) for h in header_row]
        
        if 'dws_id' not in headers:
            print(f"  WARNING: {filename} has unexpected headers, skipping")
            continue
        
        batch = []
        row_count = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i] is not None:
                    row_dict[headers[i]] = val
            
            dws_id = row_dict.get('dws_id')
            parameter = row_dict.get('parameter_name')
            if dws_id is None or parameter is None:
                continue
            
            try:
                dws_id = int(dws_id)
            except (ValueError, TypeError):
                continue
            
            # Extract year from ACTUAL sample date, not filename
            year = get_year(row_dict.get('sample_date'))
            if year is None or year < 2015 or year > 2026:
                continue
            
            # Track unique location
            if dws_id not in locations_set:
                dws_name = str(row_dict.get('dws_name') or '')
                city_hint = None
                m = re.search(r'\(([^)]+)\)\s*$', dws_name)
                if m:
                    c = m.group(1).strip()
                    if len(c) < 50 and not c.startswith('#'):
                        city_hint = c
                locations_set[dws_id] = {
                    'dws_name': dws_name,
                    'owner_name': str(row_dict.get('owner_name') or ''),
                    'regulation': str(row_dict.get('regulation') or ''),
                    'dws_category': str(row_dict.get('dws_category') or ''),
                    'phu_name': str(row_dict.get('phu_name') or ''),
                    'city_hint': city_hint,
                }
            
            # Format sample date
            sd = row_dict.get('sample_date')
            sd_str = None
            if hasattr(sd, 'strftime'):
                sd_str = sd.strftime('%Y-%m-%d')
            elif sd:
                sd_str = str(sd)[:10]
            
            batch.append((
                dws_id,
                sd_str,
                str(row_dict.get('sample_type') or ''),
                str(row_dict.get('sample_location') or ''),
                str(parameter),
                parse_result(row_dict.get('result_value')),
                str(row_dict.get('result_unit') or ''),
                str(row_dict.get('parameter_limit') or ''),
                str(row_dict.get('exceedance') or ''),
                str(row_dict.get('parameter_group') or ''),
                row_dict.get('dwis_record_id'),
                str(row_dict.get('qualifier') or ''),
                year,
            ))
            row_count += 1
            
            if len(batch) >= 50000:
                cur.executemany("""
                    INSERT INTO test_results 
                    (dws_id, sample_date, sample_type, sample_location, parameter_name, 
                     result_value, result_unit, parameter_limit, exceedance, 
                     parameter_group, dwis_record_id, qualifier, year)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, batch)
                batch = []
                total_rows += 50000
                print(f"    Inserted {total_rows} total rows so far...", end='\r')
        
        if batch:
            cur.executemany("""
                INSERT INTO test_results 
                (dws_id, sample_date, sample_type, sample_location, parameter_name, 
                 result_value, result_unit, parameter_limit, exceedance, 
                 parameter_group, dwis_record_id, qualifier, year)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, batch)
            total_rows += len(batch)
        
        print(f"    {filename}: {row_count} rows processed")
        wb.close()
    
    print(f"\nTotal rows: {total_rows}")
    
    # Insert locations
    loc_batch = []
    for dws_id, info in locations_set.items():
        loc_batch.append((
            dws_id, info['dws_name'], info['owner_name'],
            info['regulation'], info['dws_category'],
            info['phu_name'], info['city_hint']
        ))
    
    cur.executemany("""
        INSERT INTO locations (dws_id, dws_name, owner_name, regulation, dws_category, phu_name, city_hint)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, loc_batch)
    print(f"Unique locations: {len(locations_set)}")
    
    # Rebuild FTS5
    cur.execute("CREATE VIRTUAL TABLE IF NOT EXISTS locations_fts USING fts5(dws_id UNINDEXED, dws_name, owner_name, phu_name, tokenize='porter unicode61')")
    cur.execute("DELETE FROM locations_fts")
    cur.execute("INSERT INTO locations_fts (dws_id, dws_name, owner_name, phu_name) SELECT dws_id, dws_name, owner_name, phu_name FROM locations")
    
    conn.commit()
    
    # Verify
    tc = cur.execute("SELECT COUNT(*) FROM test_results").fetchone()[0]
    lc = cur.execute("SELECT COUNT(*) FROM locations").fetchone()[0]
    yrs = cur.execute("SELECT DISTINCT year FROM test_results ORDER BY year").fetchall()
    print(f"\nVerification: {lc} locations, {tc} test results")
    print(f"Years: {[y[0] for y in yrs]}")
    
    for y, in yrs:
        c = cur.execute("SELECT COUNT(*) FROM test_results WHERE year = ?", (y,)).fetchone()[0]
        print(f"  {y}: {c:,} samples")
    
    conn.close()
    print(f"\nDatabase saved to: {DB_PATH}")


if __name__ == '__main__':
    print("=" * 60)
    print("Ontario Tap Water Data Processing Pipeline")
    print("=" * 60)
    print()
    build_database()